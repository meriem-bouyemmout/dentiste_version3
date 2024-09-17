from tkinter import *
from tkinter import ttk
import customtkinter as ctk
from PIL import Image,ImageTk
import tkinter.messagebox as mb
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pathlib
import sqlite3
import os
import subprocess
from tkinter import filedialog
from openpyxl.styles import Font, PatternFill
import fdb
import pyodbc


class List_Patients:
    def __init__(self,mast):
        self.master = mast
        self.master.title("Les patients")
        ctk.set_appearance_mode("light")  # Modes: system (default), light, dark
        ctk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green
        self.width = self.master.winfo_screenwidth()
        self.height = self.master.winfo_screenheight()
        self.master.geometry("{w}x{h}+0+0".format(w=self.width,h=self.height))
        self.master.state("zoomed")
        
        
        
        #=========================university management system=======================================#
   
        self.Frameleft = ctk.CTkFrame(self.master,fg_color="#BCD2EE", width=300)
        self.Frameleft.pack(side=LEFT, fill=Y)
        ################################################################################################
       

        #############################################################################################
        self.Nom = ctk.CTkLabel(self.Frameleft,text='Nom', font=('Helvetica',15))
        self.Nom.place(x=10,y=20 )
        self.Prenom = ctk.CTkLabel(self.Frameleft,text='Prenom', font=('Helvetica',15))
        self.Prenom.place(x=10,y=60 )
        self.Age = ctk.CTkLabel(self.Frameleft,text='Age', font=('Helvetica',15))
        self.Age.place(x=10,y=100)
        # self.Motif = ctk.CTkLabel(self.Frameleft,text='Motif', font=('Helvetica',15))
        # self.Motif.place(x=10,y=140 )
        # self.Jour = ctk.CTkLabel(self.Frameleft,text='Jour', font=('Helvetica',15))
        # self.Jour.place(x=10,y=180)
        # self.Rendez_vous = ctk.CTkLabel(self.Frameleft,text='Rendez-vous', font=('Helvetica',15))
        # self.Rendez_vous.place(x=10,y=220)
        # self.Montant_total = ctk.CTkLabel(self.Frameleft,text='Montant total', font=('Helvetica',15))
        # self.Montant_total.place(x=10,y=260)
        # self.Versement = ctk.CTkLabel(self.Frameleft,text='Versement', font=('Helvetica',15))
        # self.Versement.place(x=10,y=300)
        # self.Reste = ctk.CTkLabel(self.Frameleft,text='Reste', font=('Helvetica',15))
        # self.Reste.place(x=10,y=340)
        self.Tel = ctk.CTkLabel(self.Frameleft,text='Telephone', font=('Helvetica',15))
        self.Tel.place(x=10,y=140)
        
##################################
        
        self.nom = StringVar()
        self.prenom = StringVar()
        self.age = StringVar()
        # self.motif = StringVar()
        # self.jour = StringVar()
        # self.rendez_vous = StringVar()
        # self.montant_ttl = StringVar()
        # self.versement = StringVar()
        # self.reste = StringVar()
        self.num_de_tel = StringVar()

    


    
########################################################
        self.nom_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.nom)
        self.nom_entry.configure(justify="center")
        self.nom_entry.place(x=120,y=20)
        self.prenom_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.prenom)
        self.prenom_entry.configure(justify="center")
        self.prenom_entry.place(x=120,y=60)
        self.age_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.age)
        self.age_entry.configure(justify="center")
        self.age_entry.place(x=120,y=100)
        # self.list = ["Consultation", "Vaccination", "Examen médical"]
        # self.motif_entry = ctk.CTkComboBox(self.Frameleft, font=('tahoma',12), values=self.list)
        # self.motif_entry.configure(justify="center")
        # self.motif_entry.place(x=120,y=140)
        # self.jour_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.jour)
        # self.jour_entry.configure(justify="center")
        # self.jour_entry.place(x=120,y=180)
        # self.rendez_vous_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.rendez_vous)
        # self.rendez_vous_entry.configure(justify="center")
        # self.rendez_vous_entry.place(x=120,y=220)
        # self.montant_total_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.montant_ttl)
        # self.montant_total_entry.configure(justify="center")
        # self.montant_total_entry.place(x=120,y=260)
        # self.versement_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.versement)
        # self.versement_entry.configure(justify="center")
        # self.versement_entry.place(x=120,y=300)
        # self.reste_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable =self.reste)
        # self.reste_entry.configure(justify="center")
        # self.reste_entry.place(x=120,y=340)
        self.tel_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable =self.num_de_tel)
        self.tel_entry.configure(justify="center")
        self.tel_entry.place(x=120,y=140)

        self.montants = {
            "Consultation": 2000,
            "Vaccination": 5000,
            "Examen médical": 2000
        }

        # Utiliser after pour vérifier la sélection de la combobox périodiquement
        # self.previous_motif = None
        # self.check_combobox_selection()

        # # Lier la fonction de rappel à l'événement de sélection de la Combobox
        # self.motif.trace_add("write", self.update_montant_total)

        # # Bind the calculation of reste to changes in montant_ttl and versement
        # self.montant_ttl.trace_add("write", self.update_reste)
        # self.versement.trace_add("write", self.update_reste)


        self.buttonAdd=ctk.CTkButton(self.Frameleft,text='Ajouter', command=self.ajouter,  font=('Helvetica',15,'bold'))
        self.buttonAdd.place(x=10,y=450)
        self.buttonDELETE=ctk.CTkButton(self.Frameleft,text='Supprimer', command=self.delete,  font=('Helvetica',15,'bold'))
        self.buttonDELETE.place(x=155,y=450)
        self.buttonUP=ctk.CTkButton(self.Frameleft,text='Modifier', command=self.update,  font=('Helvetica',15,'bold'))
        self.buttonUP.place(x=10,y=485)
        self.buttonRESET=ctk.CTkButton(self.Frameleft,text='Nettoyer', command=self.netoyer,  font=('Helvetica',15,'bold'))
        self.buttonRESET.place(x=155,y=485)


        image_path = "images\\download-_1_.ico"
        image = Image.open(image_path)    
        image = image.resize((30, 30))
        photo_image = ImageTk.PhotoImage(image)



        self.buttonEXCEL=ctk.CTkButton(self.Frameleft, image=photo_image, text='Exporter vers excel', command=self.export_to_excel,height=40,  font=('Helvetica',15,'bold'))
        self.buttonEXCEL.place(x=10,y=550)

        
        ####################################### RIGHT ####################################################
        self.Frameright = ctk.CTkFrame(self.master, height=800)
        self.Frameright.pack(fill=BOTH, expand=True)
        ###########################################################################################################
        


        # ##################################################################################################
        # Frame pour les champs de saisie et les boutons
        self.Framerighttop = ctk.CTkFrame(self.Frameright, fg_color="#BCD2EE", height=70)

        # Champ de saisie pour l'ID
        self.id_entry = ctk.CTkEntry(self.Framerighttop, font=('Helvetica', 18, 'bold'), width=15, placeholder_text="ID")
        self.id_entry.grid(row=0, column=0, sticky='nsew', pady=10, padx=5)

        # Champ de saisie pour le nom
        self.nom_entry_recherche = ctk.CTkEntry(self.Framerighttop, font=('Helvetica', 18, 'bold'), width=20, placeholder_text="Nom")
        self.nom_entry_recherche.grid(row=0, column=1, sticky='nsew', pady=10, padx=5)

        # Champ de saisie pour le prénom
        self.prenom_entry_recherche= ctk.CTkEntry(self.Framerighttop, font=('Helvetica', 18, 'bold'), width=20, placeholder_text="Prénom")
        self.prenom_entry_recherche.grid(row=0, column=2, sticky='nsew', pady=10, padx=5)

        # Bouton de recherche
        self.rechercher_button = ctk.CTkButton(self.Framerighttop, text='Rechercher', command=self.rechercher_ligne, font=('Helvetica', 16, 'bold'), height=30, width=90)
        self.rechercher_button.grid(row=0, column=3, sticky='nsew', pady=10, padx=5)

        # Bouton pour voir les détails
        self.voir_button = ctk.CTkButton(self.Framerighttop, text='Voir', command=self.voir, font=('Helvetica', 16, 'bold'), height=30, width=90)
        self.voir_button.grid(row=0, column=4, sticky='nsew', pady=10, padx=5)

        # Configuration des colonnes pour qu'elles s'étendent également
        self.Framerighttop.grid_columnconfigure(0, weight=1)
        self.Framerighttop.grid_columnconfigure(1, weight=1)
        self.Framerighttop.grid_columnconfigure(2, weight=1)
        self.Framerighttop.grid_columnconfigure(3, weight=1)
        self.Framerighttop.grid_columnconfigure(4, weight=1)

        self.Framerighttop.pack(fill=X)




        ##################################################################################################
        
        


        self.frameView = ctk.CTkFrame(self.Frameright, height=400)
        self.frameView.pack(fill=BOTH)

        self.scrollbar = Scrollbar(self.frameView, orient = VERTICAL)

        style1 = ttk.Style()
        style1.layout('my.treeview.layout',
                    [('Header', {'sticky':'nswe'})] +
                    [('Separator', {'sticky':'ew'})] +
                    [('Item..focus', {'sticky':'nswe'})] +
                    [('Item', {'sticky':'nswe'})]
                    )
        style1.configure("Treeview",  background="#00C957")
        style1.configure("Treeview.Item", font=("Helvetica", 12))
        style1.configure("Treeview.Heading",  font=("tahoma", 10))

        

        self.table = ttk.Treeview(self.frameView, style='Treeview.Heading', column= ("ID","Nom","Prenom","Age","Num de tel"), show='headings', height=17 , yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side=RIGHT, fill=Y)
        self.scrollbar.config(command=self.table.yview())       
        self.table.pack(fill=BOTH)
         

        self.table.heading("ID",text="ID")
        self.table.heading("Nom",text="Nom")
        self.table.heading("Prenom",text="Prenom")
        self.table.heading("Age",text="Age")
        # self.table.heading("Motif de consultation",text="Motif de consultation")
        # self.table.heading("Jour",text="Jour")
        # self.table.heading("Rendez-vous",text="Rendez-vous")
        # self.table.heading("Montant total",text="Montant total")
        # self.table.heading("Versement",text="Versement")
        # self.table.heading("Reste",text="Reste")
        self.table.heading("Num de tel",text="Num de tel")
       
        self.table.column("ID", anchor=W, width=5)
        self.table.column("Nom", anchor=W, width=5)
        self.table.column("Prenom", anchor=W, width=6)
        self.table.column("Age", anchor=W, width=6)
        # self.table.column("Motif de consultation", anchor=W, width=6)
        # self.table.column("Jour", anchor=W, width=6)
        # self.table.column("Rendez-vous", anchor=W, width=6)
        # self.table.column("Montant total", anchor=W, width=6)
        # self.table.column("Versement", anchor=W, width=6)
        # self.table.column("Reste", anchor=W, width=6)
        self.table.column("Num de tel", anchor=W, width=6)
        
        
        self.lire()
        self.table.bind("<ButtonRelease>", self.show)
        

        self.img = Image.open('images\\Teethcare Health Concept with Dental Care Tools and Dentist Instruments Stock Image - Image of health, dentistry_ 157806363.jpg')
        self.new_img = ImageTk.PhotoImage(self.img)
        self.imgDent = Label(self.Frameright, image=self.new_img)
        self.imgDent.pack( padx=0, pady =0)

    
    # def check_combobox_selection(self):
    #     current_motif = self.motif_entry.get()
    #     if current_motif != self.previous_motif:
    #         self.update_montant_total()
    #         self.previous_motif = current_motif
    #     self.master.after(100, self.check_combobox_selection)  # Vérifier toutes les 100 ms     
            

    # def update_montant_total(self):
    #     selected_motif = self.motif_entry.get()
    #     montant = self.montants.get(selected_motif, 0)
    #     self.montant_ttl.set(str(montant))



    # def update_reste(self, *args):
    #     try:
    #         montant_total = float(self.montant_ttl.get())
    #         versement = float(self.versement.get())
    #         reste = montant_total - versement
    #         self.reste.set(f"{reste:.2f}")
    #     except ValueError:
    #         self.reste.set("")



    def ajouter(self):
        nom = self.nom_entry.get()
        prenom = self.prenom_entry.get()
        age = self.age_entry.get()
        # motif = self.motif_entry.get()
        # jour = self.jour_entry.get()
        # rendez_vous = self.rendez_vous_entry.get()
        # montant_total = self.montant_total_entry.get()
        # versement = self.versement_entry.get()
        tel = self.tel_entry.get()

        # Calculer la valeur du reste
        

        def read_database_path(file_path='data_base.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()
            

        # Connexion à la base de données Firebird
        def read_fbclient_path(file_path='fb_client.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        database_path = read_database_path()
        fbclient_path = read_fbclient_path()


        # Connexion à la base de données Firebird
        conn = fdb.connect(
            dsn=database_path,
            user='SYSDBA',  # ou l'utilisateur configuré
            password='1234',
            fb_library_name = fbclient_path  # Spécifiez le chemin complet vers fbclient.dll
        )
        cursor = conn.cursor()

        if nom == '' or prenom == '':
            mb.showerror('Erreur', 'Veuillez saisir le nom et le prénom', parent=self.master)
            return

           
        else:
            req = "INSERT INTO Patient(NOM, PRENOM, AGE,  NUM_TEL) values (?, ?, ?, ?) RETURNING ID"
            val = (nom, prenom, age,  tel)
            cursor.execute(req, val)
            patient_id = cursor.fetchone()[0]
            conn.commit()


            # Message de succès avec l'ID du patient
            mb.showinfo('Succès', f'Données insérées. ID du patient : {patient_id}', parent=self.master)
            self.imprimer_informations_patient(patient_id)

            self.lire()
            self.netoyer()

        conn.close()
            
           
    def lire(self):
          
        def read_database_path(file_path='data_base.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()
            

        # Connexion à la base de données Firebird
        def read_fbclient_path(file_path='fb_client.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        database_path = read_database_path()
        fbclient_path = read_fbclient_path()    

        
        # Connexion à la base de données Firebird
        conn = fdb.connect(
            dsn=database_path,
            user='SYSDBA',  # ou l'utilisateur configuré
            password='1234',  # ou le mot de passe configuré
            charset='UTF8',  # Utilisez le charset correspondant à votre base de données
            fb_library_name = fbclient_path  # Spécifiez le chemin complet vers fbclient.dll
        )

        cursor = conn.cursor()
        

        # Fetch data from SQLite
        cursor.execute("SELECT * FROM Patient")
        data = cursor.fetchall()


        self.table.delete(*self.table.get_children())

        counter = 1  # Start from 1 or another appropriate value
        for i in data:
            self.table.insert('', 'end', iid=str(counter), values=i)
            counter += 1

        conn.close()  
              
    def show(self,ev): 
        selected_item = self.table.selection()
    
        # Extract the unique identifier (e.g., ID)
        self.row_id = self.table.item(selected_item, "values")[0]

        self.data = self.table.focus()
        alldata = self.table.item(self.data)
        print(self.row_id)
        val = alldata['values']
        self.nom.set(val[1])
        self.prenom.set(val[2])
        self.age.set(val[3])
        # self.motif_entry.set(val[4])      
        # self.jour.set(val[5])
        # self.rendez_vous.set(val[6])
        # self.montant_ttl.set(val[7])
        # self.versement.set(val[8])
        # self.reste.set(val[9])
        self.num_de_tel.set(val[4])

    def voir(self):
        # Call your read function to refresh the table with all data
        self.lire()

        # Optionally, you can clear the search entry if you have one
        self.rechercher_entry.delete(0, 'end')

    def netoyer(self):
        self.nom_entry.delete(0,'end')
        self.prenom_entry.delete(0,'end')
        self.age_entry.delete(0,'end')
        #self.motif_entry.delete(0,'end')
        # self.jour_entry.delete(0,'end')
        # self.rendez_vous_entry.delete(0,'end')
        # self.montant_total_entry.delete(0,'end')
        # self.versement_entry.delete(0,'end')
        # self.reste_entry.delete(0,'end')
        self.tel_entry.delete(0,'end')


    def delete(self):
    
        def read_database_path(file_path='data_base.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()
            

        # Connexion à la base de données Firebird
        def read_fbclient_path(file_path='fb_client.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        database_path = read_database_path()
        fbclient_path = read_fbclient_path()    

        
        # Connexion à la base de données Firebird
        conn = fdb.connect(
            dsn=database_path,
            user='SYSDBA',  # ou l'utilisateur configuré
            password='1234',  # ou le mot de passe configuré
            charset='UTF8',  # Utilisez le charset correspondant à votre base de données
            fb_library_name = fbclient_path  # Spécifiez le chemin complet vers fbclient.dll
        )
        cursor = conn.cursor()
        req = ("delete from patient where ID="+self.row_id)
        cursor.execute(req)
        conn.commit()
        conn.close()
        mb.showinfo('Supprimer', 'Le patient a été supprimé', parent=self.master)
        self.lire()
        self.netoyer() 

    def update(self):
      
        nom = self.nom_entry.get()
        prenom = self.prenom_entry.get()
        age = self.age_entry.get()
        # motif = self.motif_entry.get()
        # jour = self.jour_entry.get()
        # rendez_vous = self.rendez_vous_entry.get()
        # montant_total = self.montant_total_entry.get()
        # versement = self.versement_entry.get()
        # reste = self.reste_entry.get()
        tel = self.tel_entry.get()


        
        def read_database_path(file_path='data_base.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()
            

        # Connexion à la base de données Firebird
        def read_fbclient_path(file_path='fb_client.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        database_path = read_database_path()
        fbclient_path = read_fbclient_path()


        conn = fdb.connect(
                dsn=database_path,
                user='SYSDBA',  # ou l'utilisateur configuré
                password='1234',  # ou le mot de passe configuré
                charset='UTF8',  # Utilisez le charset correspondant à votre base de données
                fb_library_name = fbclient_path  # Spécifiez le chemin complet vers fbclient.dll
          )
        cursor = conn.cursor()

        if (nom == '' or prenom == '' ) :
           mb.showerror('Erreur','Veuiller saisir le nom et le prenom', parent=self.master)
        else :
            req = "UPDATE PATIENT set ID=? ,NOM=?, PRENOM=?, AGE=?, NUM_TEL=? WHERE ID=? "  
            val = (self.row_id, nom, prenom, age,  tel, self.row_id)          
            cursor.execute(req, val)        
            
            mb.showinfo('Mise a jour','Le patient a été modifier', parent=self.master)
            
            # Fetch data from SQLite
            req2 = "SELECT * FROM Patient where ID=?"
            cursor.execute(req2,(self.row_id,))
            data = cursor.fetchall()
            
    
            self.table.delete(*self.table.get_children())

            counter = 1  # Start from 1 or another appropriate value
            for i in data:
                self.table.insert('', 'end', iid=str(counter), values=i)
                counter += 1
            conn.commit()
            conn.close()  
            
                    
            self.netoyer()
        

    def rechercher_ligne(self):


        # Récupérer les valeurs des champs de saisie
        id_value = self.id_entry.get()
        nom_value = self.nom_entry_recherche.get()
        prenom_value = self.prenom_entry_recherche.get()

        def read_database_path(file_path='data_base.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        # Connexion à la base de données Firebird
        def read_fbclient_path(file_path='fb_client.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        database_path = read_database_path()
        fbclient_path = read_fbclient_path()

        conn = fdb.connect(
            dsn=database_path,
            user='SYSDBA',  # ou l'utilisateur configuré
            password='1234',  # ou le mot de passe configuré
            charset='UTF8',  # Utilisez le charset correspondant à votre base de données
            fb_library_name=fbclient_path  # Spécifiez le chemin complet vers fbclient.dll
        )
        cursor = conn.cursor()
        
        # Déterminer quelle recherche effectuer
        
        if (id_value != '' and nom_value == '' and prenom_value == '') :
            # Rechercher par ID
            req = "SELECT * FROM PATIENT WHERE ID LIKE ?"
            cursor.execute(req, ('%' + id_value + '%',))
        else:
            if (id_value == '' and nom_value != '' and prenom_value == '') :
                # Rechercher par Nom
                req = "SELECT * FROM PATIENT WHERE NOM LIKE ?"
                cursor.execute(req, ('%' + nom_value + '%',))
            else:
                if (id_value == '' and nom_value == '' and prenom_value != '') :
                    # Rechercher par Prénom
                    req = "SELECT * FROM PATIENT WHERE PRENOM LIKE ?"
                    cursor.execute(req, ('%' + prenom_value + '%',))
                else :
                    if ( id_value == '' and nom_value == '' and prenom_value == '' ):
                       mb.showerror("Erreur", "Veuillez saisir une valeur dans un champ", parent=self.master)
        
            

        resultats = cursor.fetchall()

        if not resultats:
            mb.showerror("Erreur", "Le patient n'existe pas", parent=self.master)
        else:
            # Effacer les anciennes entrées dans le tableau
            for row in self.table.get_children():
                self.table.delete(row)

            # Afficher les résultats dans le tableau
            for resultat in resultats:
                self.table.insert("", "end", values=resultat)

        conn.close()


      

    def export_to_excel(self):
        # Connect to SQLite database
        def read_database_path(file_path='data_base.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()
            

        # Connexion à la base de données Firebird
        def read_fbclient_path(file_path='fb_client.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        database_path = read_database_path()
        fbclient_path = read_fbclient_path()


        conn = fdb.connect(
                dsn=database_path,
                user='SYSDBA',  # ou l'utilisateur configuré
                password='1234',  # ou le mot de passe configuré
                charset='UTF8',  # Utilisez le charset correspondant à votre base de données
                fb_library_name = fbclient_path  # Spécifiez le chemin complet vers fbclient.dll
          )
        cursor = conn.cursor()

        # Fetch data from SQLite
        cursor.execute("SELECT * FROM PATIENT")
        data = cursor.fetchall()

        # Create a new Excel workbook and sheet
        wb = Workbook()
        ws = wb.active

        # Define styles for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        # Write headers to Excel sheet
        headers = [description[0] for description in cursor.description]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Define a list of colors for the columns
        colors = ["FFC7CE", "C6EFCE", "FFEB9C", "9CC3E6", "D9EAD3", "EAD1DC", "FFF2CC", "D9D2E9", "F4CCCC", "D0E0E3"]

        # Write data to Excel sheet and apply column colors
        for row_num, row in enumerate(data, 2):  # Start from row 2 because row 1 is the header
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                column_fill = PatternFill(start_color=colors[(col_num - 1) % len(colors)], end_color=colors[(col_num - 1) % len(colors)], fill_type="solid")
                cell.fill = column_fill

        # Save Excel file
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb.save(file_path)
            mb.showinfo("Export Successful", f"Data exported to {file_path}", parent=self.master)
        
        conn.close()

    def imprimer_informations_patient(self, patient_id):

        
        def read_database_path(file_path='data_base.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()
            

        # Connexion à la base de données Firebird
        def read_fbclient_path(file_path='fb_client.txt'):
            with open(file_path, 'r') as file:
                return file.read().strip()

        database_path = read_database_path()
        fbclient_path = read_fbclient_path()


        conn = fdb.connect(
                dsn=database_path,
                user='SYSDBA',  # ou l'utilisateur configuré
                password='1234',  # ou le mot de passe configuré
                charset='UTF8',  # Utilisez le charset correspondant à votre base de données
                fb_library_name = fbclient_path  # Spécifiez le chemin complet vers fbclient.dll
          )
        cursor = conn.cursor()

        try:
            req = "SELECT * FROM Patient WHERE ID = ?"
            cursor.execute(req, (patient_id,))
            patient_info = cursor.fetchone()
            
            if patient_info:
                # Créer le contenu à imprimer
                contenu = f"""
                Informations du patient :
                ID : {patient_info[0]}
                Nom : {patient_info[1]}
                Prénom : {patient_info[2]}
                Âge : {patient_info[3]}
                Téléphone : {patient_info[4]}
                """
                
                if os.name == 'nt':  # Pour Windows
                    # Imprimer en utilisant notepad
                    temp_file_path = 'temp_print_file.txt'
                    with open(temp_file_path, 'w', encoding='utf-8') as temp_file:
                        temp_file.write(contenu)
                    subprocess.run(['notepad', '/p', temp_file_path])
                    os.remove(temp_file_path)
                else:  # Pour Unix (Linux, MacOS)
                    # Imprimer en utilisant lp
                    subprocess.run(['lp'], input=contenu.encode('utf-8'))
                
            else:
                print(f"Aucune information trouvée pour l'ID {patient_id}.")
        
        except sqlite3.Error as e:
            print(f"Erreur lors de la récupération des informations du patient : {e}")
        
        finally:
            conn.close()

if (__name__ == '__main__'):
    window = ctk.CTk()
    window.iconbitmap('images\\download.ico')
    std = List_Patients(window)
    mainloop()